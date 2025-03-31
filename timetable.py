import openpyxl 
wb_obj = openpyxl.load_workbook('timetable.xlsx')
sheet_obj = wb_obj.active
# print(wb_obj.sheetnames)
row = sheet_obj.max_row
col = sheet_obj.max_column


# print(sheet_obj)
# print(row , col)

# for i in range(2,19):
#     for j in range(1,col+1):
#         cell=sheet_obj.cell(row=i,column=j)
#         if( cell.value==None  or len(cell.value)!=0 ):
#             print(cell.value, end=' ')
#     print('\n')






data= sheet_obj.cell(row=4,column=2)    
if (len(data.value)==0):
    print(data.value)
    print("hey : ")
else :
    print("data is empty")
    print(data.value)

# cell=sheet_obj.cell(row=11,column=1)

# print(cell.value)
class_info={'Lecture' : [] ,'Batch': [], 'Faculty' : [], 'Classroom': []}



cell=sheet_obj.cell(row=4,column=2) # it is string returned
val=cell.value
val2=val.split() # it will convert into list
print(val2)

class_info['Lecture'].append(val2[0])
class_info['Batch'].append(val2[1])
class_info['Faculty'].append(val2[2])
class_info['Classroom'].append(val2[3])
# class_info['Lecture'].append('IT')


print(class_info)


# print(type(val))
