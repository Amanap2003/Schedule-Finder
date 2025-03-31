# import pandas as pd
import openpyxl
import openpyxl as op

w_object=openpyxl.load_workbook('timetable.xlsx') # it is our workbook object
s_obj = w_object.active # it is our sheet object

row = s_obj.max_row
col = s_obj.max_column


cell_merge=s_obj.merged_cells
# print(cell_merge.row_end)
# print(type(cell_merge))
l=list(cell_merge)
z="ABC".replace("",'=')
print(z)
# print(type(z))
print("\n")
# print(int("ABc23"))
# print(l)
# string=str(l[0])
# print("\n")
list_merge_cells=[]
for i in range(len(l)):
    list_merge_cells.append(str(l[i]))
print(list_merge_cells)

print("\n")

day_cells ={} # it will store the days and their respective cells that are merged <>
for k in range(len(list_merge_cells)) :
    # for l in range(2):
        list_merge_cells[k]=list_merge_cells[k].split(':')
list_merge_cells.sort()


print("After convertin list of merged cells is : \n\n",list_merge_cells)

for i in range(len(list_merge_cells)) :
      for k in range (2):
            # print(i[k])
            list_merge_cells[i][k]=list_merge_cells[i][k].replace("A",'')
      
print("Now after removing the characters in list , our final list is : \n",list_merge_cells)


for h in range(0,len(list_merge_cells)):
    if (not list_merge_cells[h][0].isdigit() or not list_merge_cells[h][1].isdigit()) :
          # print("POP : ",h,l)
        list_merge_cells.pop(h)
        break 

    
      

for i in range(0,len(list_merge_cells)) :
      for k in range (2):
            # print(i[k])
            list_merge_cells[i][k]=int(list_merge_cells[i][k])
print(list_merge_cells)


               
               
# list_merge_cells.pop(0)


list_merge_cells.sort()
print(list_merge_cells)

list_range_day=[]

for i in range(len(list_merge_cells)):
      list_range_day.append(list_merge_cells[i][1]-list_merge_cells[i][0]+1)
print(list_range_day)




    











# print("string is : ",string)
# l2=string.split(':')
# print("list of cells merge is : ",l2)

# print(l)
# print("\n")
# print(type(cell_merge))
# print("\n")

# for i in cell_merge :
#     print(i)





# l=[]

# for i in cell_merge :
#     # print(cell_merge)
#     print(i)
#     l.append(i)

# l2=[]
# cell=s_obj.cell(row=2,column=1).


# for j in l :
#     # l2=j.split(':')
#     print(j,end=' ')
# # print(l2)
# print(type(l[0]))



# for i in range(1,10):
#     cell=s_obj.cell(row=i,column=i)
#     print(cell.value)

# cell=s_obj.cell(row=3,column=2)  # create the object for accessing the cell value

# for i in range(1,s_obj.max_row) :
#     cell=s_obj.cell(row=i,column=1)
#     print(cell.value)
print("Max rows are : ",s_obj.max_row)