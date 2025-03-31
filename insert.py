# import main
import openpyxl
import pymongo
# from GUI import faculty


# attributes = {
#             'Lecture' : ''  ,
#             'Lecture_Type': '',   
#             'Batch': ''     , 
#             'Faculty': ''   , 
#             'Room_No': ''
#             }

# def connect():
#     client=pymongo.MongoClient("mongodb://localhost:27017")
#     print(client)
#     DATABASE ='Timetable'
#     db=client['PythonProject']
#     collection = db[DATABASE]


wb_obj = openpyxl.load_workbook('TT_new.xlsx')
# wb_obj = openpyxl.load_workbook('Book2.xlsx')
sheet_obj = wb_obj.active

row = sheet_obj.max_row
# col = sheet_obj.max_column
col=10

cell_merge=sheet_obj.merged_cells
# print(cell_merge.row_end)
# print(type(cell_merge))
l=list(cell_merge)

# print(z)
# print(type(z))
# print("\n")
# print(int("ABc23"))
# print(l)
# string=str(l[0])
# print("\n")
list_merge_cells=[]
for i in range(len(l)):
    list_merge_cells.append(str(l[i]))
# print(list_merge_cells)

# print("\n")

day_cells ={} # it will store the days and their respective cells that are merged <>
for k in range(len(list_merge_cells)) :
    # for l in range(2):
        list_merge_cells[k]=list_merge_cells[k].split(':')
list_merge_cells.sort()


# print("After convertin list of merged cells is : \n\n",list_merge_cells)

for i in range(len(list_merge_cells)) :
      for k in range (2):
            # print(i[k])
            list_merge_cells[i][k]=list_merge_cells[i][k].replace("A",'')
      
# print("Now after removing the characters in list , our final list is : \n",list_merge_cells)

for h in range(0,len(list_merge_cells)):
    if (not list_merge_cells[h][0].isdigit() or not list_merge_cells[h][1].isdigit()) :
          # print("POP : ",h,l)
        list_merge_cells.pop(h)
        break 

for i in range(0,len(list_merge_cells)) :
      for k in range (2):
            # print(i[k])
            list_merge_cells[i][k]=int(list_merge_cells[i][k])
# print(list_merge_cells)
# list_merge_cells.pop(0)
list_merge_cells.sort()
# print(list_merge_cells)

list_range_day=[]

for i in range(len(list_merge_cells)):
      list_range_day.append(list_merge_cells[i][1]-list_merge_cells[i][0]+1)

teachers=  []

# print(list_range_day)

list_day=['MON','TUE','WED','THU','FRI','SAT']
#------------------------------------------------------------------------------
def add_into_db(collection, faculty, batches, Ltype , Lcode) :
    print(type(collection))
    # collection = db[DATABASE]
    attributes={}
    D=0
    count=1

    
    for i in range (3,row+1):
           
        for j in range(2,col+1):
# or we can redefine dict here .....><>
            cell=sheet_obj.cell(row=i,column=j)
            if (len(cell.value)!=0):

                
                val=cell.value.split()

                Batches=val[1].split(',')
                for k in range(len(Batches)):
                    
                    attributes['Day']=list_day[D]
                    
                    attributes['Lecture_Type']=val[0][0]
                    attributes['Lecture']=val[0]
                    
                    attributes['Batch']=Batches[k]
                    attributes['Faculty']=val[2]
                    # my_set = set(faculty)
                    # my_set.add(val[2])
                    # faculty = list(my_set)
                    # faculty=set(faculty)
                    faculty.add(val[2])
                    batches.add(Batches[k])
                    Ltype.add(val[0][0])
              
                    Lcode.add(val[0])

                    # faculty.append(val[2])
                    attributes['Room_No']=val[3]
                    attributes['Time']=sheet_obj.cell(row=2,column=j).value
                    # print(teachers) 
                    # print(attributes)
                    collection.insert_one(attributes)
                    # we have to redefine the dictionary to avoid duplicate values of ID 
                    attributes = {}
                    # 'Day`': '',
                    # 'Lecture' : ''  ,
                    # 'Lecture_Type': '',   
                    # 'Batch': ''     , 
                    # 'Faculty': ''   , 
                    # 'Room_No': '' ,
                    # 'Time': ''
                    # }
        # to calculate the day from the merged cell range <>
        count=count+1
        if(count > list_range_day[D]):
            D=D+1
            count=0 

     
    